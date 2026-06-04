"""Small XLSX table writer shared by manufacturing output commands."""

from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def write_xlsx_table(
    output_file: Path,
    *,
    columns: Sequence[str],
    rows: Sequence[Mapping[str, object]],
    sheet_name: str,
    highlighted_rows: Sequence[bool] = (),
) -> None:
    """Write a single-sheet XLSX workbook from named rows.

    All values are written as strings with Excel's text number format so values
    such as ``0603`` stay visibly unchanged.
    """
    output_file.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = cast(Worksheet, workbook.active)
    worksheet.title = _safe_sheet_name(sheet_name)
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill(
        start_color="000000",
        end_color="000000",
        fill_type="solid",
    )
    dnp_fill = PatternFill(
        start_color="FFF2CC",
        end_color="FFF2CC",
        fill_type="solid",
    )
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_font = Font(color="FFFFFF", bold=True)
    text_alignment = Alignment(horizontal="left", vertical="center")

    for column_index, column in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=str(column))
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = text_alignment
        cell.number_format = "@"

    for row_index, row in enumerate(rows, start=2):
        highlighted = _row_highlighted(highlighted_rows, row_index - 2)
        for column_index, column in enumerate(columns, start=1):
            value = _sanitize_cell_value(row.get(column, ""))
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.border = thin_border
            cell.alignment = text_alignment
            cell.number_format = "@"
            if highlighted:
                cell.fill = dnp_fill

    _size_columns(worksheet, len(columns), len(rows) + 1)
    workbook.save(output_file)


def _sanitize_cell_value(value: object) -> str:
    """Return a text-safe value for openpyxl cells."""
    text = "" if value is None else str(value)
    return _ILLEGAL_XML_RE.sub("", text)


def _safe_sheet_name(sheet_name: str) -> str:
    """Return an Excel-safe sheet name."""
    safe = re.sub(r"[\[\]:*?/\\]", "_", sheet_name or "Sheet1").strip()
    return (safe or "Sheet1")[:31]


def _row_highlighted(highlighted_rows: Sequence[bool], zero_based_index: int) -> bool:
    """Return whether a data row should use highlight styling."""
    return zero_based_index < len(highlighted_rows) and highlighted_rows[
        zero_based_index
    ]


def _size_columns(worksheet: Worksheet, column_count: int, row_count: int) -> None:
    """Apply simple width sizing without a larger formatting dependency."""
    for column_index in range(1, column_count + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row_index in range(1, row_count + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            max_length = max(max_length, len(str(value or "")))
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 8),
            60,
        )
